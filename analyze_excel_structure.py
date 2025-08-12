#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analyze the complete Excel file structure to find model information
分析完整的Excel檔案結構以尋找模型資訊
"""

import pandas as pd
import openpyxl
import os

def analyze_excel_structure(filename: str):
    """分析Excel檔案的完整結構"""
    
    if not os.path.exists(filename):
        print(f"❌ 檔案不存在: {filename}")
        return
    
    print("=" * 60)
    print(f"分析Excel檔案結構: {filename}")
    print("=" * 60)
    
    try:
        # 使用openpyxl讀取原始結構
        workbook = openpyxl.load_workbook(filename)
        
        print(f"📋 工作表數量: {len(workbook.sheetnames)}")
        print(f"📋 工作表名稱: {workbook.sheetnames}")
        
        for sheet_name in workbook.sheetnames:
            print(f"\n📊 工作表: {sheet_name}")
            worksheet = workbook[sheet_name]
            
            print(f"  最大行數: {worksheet.max_row}")
            print(f"  最大列數: {worksheet.max_column}")
            
            # 檢查前10行的所有內容
            print(f"  前10行內容:")
            for row in range(1, min(11, worksheet.max_row + 1)):
                row_data = []
                for col in range(1, min(15, worksheet.max_column + 1)):  # 只檢查前15列
                    cell = worksheet.cell(row=row, column=col)
                    value = str(cell.value) if cell.value is not None else '[空]'
                    row_data.append(value[:15])  # 限制長度
                print(f"    行{row}: {row_data}")
            
            # 搜尋包含模型相關關鍵字的儲存格
            print(f"\n🔍 搜尋模型相關資訊:")
            model_keywords = ['gemini', 'gemma', 'chatgpt', 'gpt', 'claude', 'model', '模型', 'ai', '人工智慧']
            found_model_info = []
            
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        cell_text = str(cell.value).lower()
                        for keyword in model_keywords:
                            if keyword in cell_text:
                                found_model_info.append({
                                    'row': row,
                                    'col': col,
                                    'value': str(cell.value),
                                    'keyword': keyword
                                })
            
            if found_model_info:
                print(f"  找到 {len(found_model_info)} 個相關資訊:")
                for info in found_model_info[:10]:  # 只顯示前10個
                    print(f"    行{info['row']}列{info['col']}: '{info['value']}' (關鍵字: {info['keyword']})")
            else:
                print(f"  未找到模型相關資訊")
        
        # 使用pandas讀取不同的header設定
        print(f"\n📊 使用不同header設定讀取:")
        for header_row in [None, 0, 1, 2, 3]:
            try:
                df = pd.read_excel(filename, header=header_row)
                print(f"  header={header_row}: {len(df)} 行 x {len(df.columns)} 欄")
                if len(df.columns) > 0:
                    print(f"    欄位: {list(df.columns)[:5]}...")  # 只顯示前5個欄位
            except Exception as e:
                print(f"  header={header_row}: 讀取失敗 - {e}")
                
    except Exception as e:
        print(f"❌ 分析過程中發生錯誤: {e}")

def main():
    """主函數"""
    
    test_files = [
        "data/[gemini2.5pro]身心障礙手冊_AI測試結果資料.xlsx",
        "data/[gemma3]身心障礙手冊_AI測試結果資料.xlsx"
    ]
    
    for test_file in test_files:
        analyze_excel_structure(test_file)
        print("\n" + "="*60 + "\n")

if __name__ == "__main__":
    main()
