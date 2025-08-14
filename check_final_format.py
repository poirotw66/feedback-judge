#!/usr/bin/env python3
import sys
import os
sys.path.append('/Users/cfh00896102/Github/feedback-judge/.venv/lib/python3.12/site-packages')

import openpyxl

# 讀取新格式的Excel結果檔案
workbook = openpyxl.load_workbook("test_with_valueset.xlsx")

print("🎯 外來函文評估結果分析（新格式）")
print("=" * 50)
print(f"📊 工作表總數: {len(workbook.sheetnames)}")
print(f"📋 工作表名稱: {workbook.sheetnames}")
print()

# 檢查每個工作表的第一行格式
for sheet_name in workbook.sheetnames:
    ws = workbook[sheet_name]
    print(f"📄 工作表: {sheet_name}")
    print(f"   🔢 資料範圍: {ws.calculate_dimension()}")
    
    # 檢查A1-D1的內容
    print("   📝 第一行（A1-D1）內容:")
    print(f"      A1: {ws['A1'].value}")
    print(f"      B1: {ws['B1'].value}")
    print(f"      C1: {ws['C1'].value}")
    print(f"      D1: {ws['D1'].value}")
    
    # 檢查第二行表頭
    print("   📝 第二行（表頭）內容:")
    print(f"      A2: {ws['A2'].value}")
    print(f"      B2: {ws['B2'].value}")
    print(f"      C2: {ws['C2'].value}")
    print(f"      D2: {ws['D2'].value}")
    print(f"      E2: {ws['E2'].value}")
    print()

print("✅ 格式檢查完成！")
print()
print("🎉 新格式特點：")
print("   • 移除了總覽分頁")
print("   • A1: '模型', B1: 模型名稱")
print("   • C1: 'valueSetId', D1: valueSetId的值")
print("   • 保持原有的案件評估結構")
print("   • 每個模型一個獨立分頁")
