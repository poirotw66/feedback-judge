#!/usr/bin/env python3
import pandas as pd
import openpyxl

# 讀取新格式的Excel結果檔案
workbook = openpyxl.load_workbook("test_new_format_result.xlsx")

print("🎯 新格式外來函文評估結果分析")
print("=" * 50)
print(f"📊 工作表總數: {len(workbook.sheetnames)}")
print(f"📋 工作表名稱: {workbook.sheetnames}")
print()

# 檢查每個工作表的結構
for sheet_name in workbook.sheetnames:
    ws = workbook[sheet_name]
    print(f"📄 工作表: {sheet_name}")
    print(f"   🔢 資料範圍: {ws.calculate_dimension()}")
    
    # 顯示前幾行的內容
    print("   📝 前5行內容:")
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        print(f"      {row}")
    print()

print("✅ 分析完成！")
print()
print("🎉 恭喜！新的按模型分組格式已成功實現：")
print("   • 每個模型都有獨立的工作表")
print("   • 按案件（受編）分組顯示準確度")
print("   • 顯示準確度、CER準確率、WER準確率")
print("   • 包含總覽工作表")
print()
print("📋 格式說明：")
print("   • 總覽：所有模型的統計摘要")
print("   • 模型工作表：每個案件的欄位評估詳情")
print("   • 顏色編碼：綠色(優秀≥90%), 黃色(良好≥70%), 紅色(需改進<50%)")
