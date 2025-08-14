#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
外來函文Excel報告生成器
Document Excel Generator
"""

import pandas as pd
import numpy as np
import io
from typing import Dict, List, Any, Tuple
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

class DocumentExcelGenerator:
    """外來函文Excel報告生成器"""
    
    def __init__(self):
        """初始化生成器"""
        self.title_font = Font(name='Arial', size=14, bold=True)
        self.header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        self.content_font = Font(name='Arial', size=10)
        
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.correct_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        self.incorrect_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    async def generate_document_evaluation_report(
        self,
        original_data: pd.DataFrame,
        evaluation_results: Dict[str, Any],
        original_filename: str
    ) -> bytes:
        """
        生成外來函文評估報告（新格式）
        
        Args:
            original_data: 原始資料
            evaluation_results: 評估結果
            original_filename: 原始檔案名稱
            
        Returns:
            bytes: Excel檔案內容
        """
        try:
            logger.info("開始生成外來函文評估報告")
            
            # 創建工作簿
            wb = Workbook()
            
            # 移除預設工作表
            wb.remove(wb.active)
            
            # 創建總覽工作表
            self._create_document_summary_sheet(wb, evaluation_results, original_filename)
            
            # 為每個欄位創建詳細工作表
            field_evaluations = evaluation_results.get('field_evaluations', {})
            for field_name, field_result in field_evaluations.items():
                self._create_field_detail_sheet(wb, field_name, field_result)
            
            # 創建原始資料工作表
            self._create_original_data_sheet(wb, original_data, original_filename)
            
            # 保存到記憶體
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info("外來函文評估報告生成完成")
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"生成外來函文評估報告時發生錯誤: {str(e)}")
            raise
    
    def _create_document_summary_sheet(
        self,
        wb: Workbook,
        evaluation_results: Dict[str, Any],
        original_filename: str
    ):
        """創建外來函文總覽工作表"""
        ws = wb.create_sheet("評估總覽", 0)
        
        # 標題
        ws['A1'] = "外來函文AI測試結果準確度評估報告"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # 基本資訊
        summary = evaluation_results.get('summary', {})
        ws['A3'] = "原始檔案:"
        ws['B3'] = original_filename
        ws['A4'] = "評估時間:"
        ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A5'] = "評估欄位數:"
        ws['B5'] = summary.get('total_fields', 0)
        ws['A6'] = "總案件數:"
        ws['B6'] = summary.get('total_cases', 0)
        ws['A7'] = "整體準確率:"
        ws['B7'] = f"{summary.get('overall_accuracy', 0):.2%}"
        
        # 表頭
        headers = ['欄位名稱', '模型名稱', '總案件數', '正確案件數', '準確率', '評估狀態']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # 填充資料
        row = 10
        field_evaluations = evaluation_results.get('field_evaluations', {})
        for field_name, field_result in field_evaluations.items():
            # 提取欄位名稱（移除模型後綴）
            display_name = field_name.split('_')[0] if '_' in field_name else field_name
            
            ws.cell(row=row, column=1, value=display_name)
            ws.cell(row=row, column=2, value=field_result.get('model_name', ''))
            ws.cell(row=row, column=3, value=field_result.get('total_cases', 0))
            ws.cell(row=row, column=4, value=field_result.get('correct_cases', 0))
            
            accuracy_rate = field_result.get('accuracy_rate', 0)
            accuracy_cell = ws.cell(row=row, column=5, value=f"{accuracy_rate:.2%}")
            
            # 根據準確率設置顏色
            if accuracy_rate >= 0.9:
                accuracy_cell.fill = self.correct_fill
                status = "優秀"
            elif accuracy_rate >= 0.7:
                accuracy_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                status = "良好"
            else:
                accuracy_cell.fill = self.incorrect_fill
                status = "需改進"
            
            ws.cell(row=row, column=6, value=status)
            
            # 設置邊框
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = self.center_alignment
            
            row += 1
        
        # 調整欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
    
    def _create_field_detail_sheet(
        self,
        wb: Workbook,
        field_name: str,
        field_result: Dict[str, Any]
    ):
        """創建欄位詳細評估工作表"""
        # 清理工作表名稱
        sheet_name = f"欄位_{field_name}"[:31]
        sheet_name = ''.join(c for c in sheet_name if c.isalnum() or c in ' _-')
        
        ws = wb.create_sheet(sheet_name)
        
        # 標題
        ws['A1'] = f"欄位 {field_name} 詳細評估結果"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:H1')
        
        # 欄位統計
        ws['A3'] = "欄位統計:"
        ws['A4'] = f"模型名稱: {field_result.get('model_name', '')}"
        ws['A5'] = f"總案件數: {field_result.get('total_cases', 0)}"
        ws['A6'] = f"正確案件數: {field_result.get('correct_cases', 0)}"
        ws['A7'] = f"準確率: {field_result.get('accuracy_rate', 0):.2%}"
        
        # 詳細結果表頭
        headers = ['案件ID', 'AI預測值', '人工正確值', '相似度', 'CER', 'WER', '匹配狀態']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # 填充案件結果
        row = 10
        case_results = field_result.get('case_results', [])
        for case_result in case_results:
            ws.cell(row=row, column=1, value=case_result.get('case_id', ''))
            ws.cell(row=row, column=2, value=case_result.get('ai_value', ''))
            ws.cell(row=row, column=3, value=case_result.get('human_value', ''))
            ws.cell(row=row, column=4, value=f"{case_result.get('similarity', 0):.2%}")
            ws.cell(row=row, column=5, value=f"{case_result.get('cer', 0):.2%}")
            ws.cell(row=row, column=6, value=f"{case_result.get('wer', 0):.2%}")
            
            # 匹配狀態
            is_correct = case_result.get('is_correct', False)
            status_cell = ws.cell(row=row, column=7, value="✓" if is_correct else "✗")
            if is_correct:
                status_cell.fill = self.correct_fill
            else:
                status_cell.fill = self.incorrect_fill
            
            # 設置邊框和對齊
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.thin_border
                if col in [4, 5, 6, 7]:  # 數值欄位置中
                    ws.cell(row=row, column=col).alignment = self.center_alignment
            
            row += 1
        
        # 調整欄寬
        column_widths = [15, 20, 20, 10, 8, 8, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
    
    def _create_summary_sheet(
        self,
        wb: Workbook,
        model_mappings: Dict[str, Dict[str, int]],
        evaluation_results: Dict[str, Any],
        original_filename: str
    ):
        """創建總覽工作表"""
        ws = wb.create_sheet("評估總覽", 0)
        
        # 標題
        ws['A1'] = "外來函文AI測試結果準確度評估報告"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # 基本資訊
        ws['A3'] = "原始檔案:"
        ws['B3'] = original_filename
        ws['A4'] = "評估時間:"
        ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A5'] = "評估模型數:"
        ws['B5'] = len(model_mappings)
        
        # 表頭
        headers = ['模型名稱', '總欄位數', '正確欄位數', '準確率', '評估狀態']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # 填充資料
        row = 8
        for model_name, eval_result in evaluation_results.items():
            ws.cell(row=row, column=1, value=model_name)
            ws.cell(row=row, column=2, value=eval_result['total_fields'])
            ws.cell(row=row, column=3, value=eval_result['correct_fields'])
            
            accuracy_cell = ws.cell(row=row, column=4, value=f"{eval_result['accuracy_rate']:.2%}")
            
            # 根據準確率設置顏色
            if eval_result['accuracy_rate'] >= 0.9:
                accuracy_cell.fill = self.correct_fill
                status = "優秀"
            elif eval_result['accuracy_rate'] >= 0.7:
                accuracy_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                status = "良好"
            else:
                accuracy_cell.fill = self.incorrect_fill
                status = "需改進"
            
            ws.cell(row=row, column=5, value=status)
            
            # 設置邊框
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = self.center_alignment
            
            row += 1
        
        # 調整欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
    
    def _create_model_detail_sheet(
        self,
        wb: Workbook,
        model_name: str,
        eval_result: Dict[str, Any],
        field_mappings: Dict[str, Tuple[int, str]],
        original_data: pd.DataFrame,
        model_mappings: Dict[str, Dict[str, int]]
    ):
        """創建模型詳細評估工作表"""
        # 清理工作表名稱（移除特殊字符）
        sheet_name = f"模型_{model_name}"[:31]  # Excel工作表名稱限制31字符
        sheet_name = ''.join(c for c in sheet_name if c.isalnum() or c in ' _-')
        
        ws = wb.create_sheet(sheet_name)
        
        # 標題
        ws['A1'] = f"模型 {model_name} 詳細評估結果"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:H1')
        
        # 模型統計
        ws['A3'] = "模型統計:"
        ws['A4'] = f"總欄位數: {eval_result['total_fields']}"
        ws['A5'] = f"正確欄位數: {eval_result['correct_fields']}"
        ws['A6'] = f"準確率: {eval_result['accuracy_rate']:.2%}"
        
        # 詳細結果表頭
        headers = ['欄位名稱', '正確值', 'AI預測值', '相似度', 'CER', 'WER', '匹配狀態', '錯誤描述']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # 填充詳細結果
        row = 9
        for field_name, field_result in eval_result['field_results'].items():
            ws.cell(row=row, column=1, value=field_name)
            ws.cell(row=row, column=2, value=field_result['correct_value'])
            ws.cell(row=row, column=3, value=field_result['predicted_value'])
            ws.cell(row=row, column=4, value=f"{field_result['similarity']:.2%}")
            ws.cell(row=row, column=5, value=f"{field_result['cer']:.2%}")
            ws.cell(row=row, column=6, value=f"{field_result['wer']:.2%}")
            
            # 匹配狀態
            status_cell = ws.cell(row=row, column=7, value="✓" if field_result['is_exact_match'] else "✗")
            if field_result['is_exact_match']:
                status_cell.fill = self.correct_fill
            else:
                status_cell.fill = self.incorrect_fill
            
            ws.cell(row=row, column=8, value=field_result['error_description'])
            
            # 設置邊框和對齊
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.thin_border
                if col in [4, 5, 6, 7]:  # 數值欄位置中
                    ws.cell(row=row, column=col).alignment = self.center_alignment
            
            row += 1
        
        # 調整欄寬
        column_widths = [15, 20, 20, 10, 8, 8, 10, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
    
    def _create_original_data_sheet(self, wb: Workbook, original_data: pd.DataFrame, original_filename: str):
        """創建原始資料工作表"""
        ws = wb.create_sheet("原始資料")
        
        # 標題
        ws['A1'] = f"原始資料 - {original_filename}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        # 添加原始資料
        for r_idx, row in enumerate(dataframe_to_rows(original_data, index=False, header=False), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.thin_border
                
                # 第一行（模型名稱）設置特殊格式
                if r_idx == 3:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
        
        # 調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            # 找到第一個非合併儲存格來獲取欄位字母
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            if column_letter:
                # 計算最大長度
                for cell in column:
                    try:
                        if hasattr(cell, 'value') and cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except Exception:
                        pass
                
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
